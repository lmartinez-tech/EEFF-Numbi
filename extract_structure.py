import json
import os
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(os.environ.get('EEFF_DATA_DIR', 'data'))
GUIDE = Path(os.environ.get('EEFF_GUIDE_PATH', DATA_DIR / 'guia_eeff.xlsx'))
SOURCES = [GUIDE] + [
    DATA_DIR / name for name in (
        '1. ENERO Balance de prueba por tercero.xlsx',
        '2. FEBRERO Balance de prueba por tercero.xlsx',
        '3. MARZO Balance de prueba por tercero.xlsx',
        '4. ABRIL Balance de prueba por tercero.xlsx',
        '5. MAYO Balance de prueba por tercero.xlsx',
        'Balance de prueba por tercero 2025.xlsx',
    )
]

def safe(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)

def row_values(ws, r, max_cols=20):
    vals = [safe(ws.cell(r, c).value) for c in range(1, min(ws.max_column, max_cols) + 1)]
    while vals and vals[-1] is None:
        vals.pop()
    return vals

def summarize(path):
    wb = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    result = {
        'path': str(path),
        'defined_names': [str(x) for x in wb.defined_names.values()],
        'sheets': [],
    }
    for ws in wb.worksheets:
        nonempty_rows = []
        formula_cells = []
        external_formula_cells = []
        for row in ws.iter_rows():
            vals = [cell.value for cell in row]
            if any(v is not None for v in vals):
                if len(nonempty_rows) < 15:
                    nonempty_rows.append({'row': row[0].row, 'values': [safe(v) for v in vals[:20]]})
                for cell in row:
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        if len(formula_cells) < 80:
                            formula_cells.append({'cell': cell.coordinate, 'formula': cell.value})
                        if '[' in str(cell.value) and len(external_formula_cells) < 50:
                            external_formula_cells.append({'cell': cell.coordinate, 'formula': cell.value})
        tail = []
        for r in range(max(1, ws.max_row - 9), ws.max_row + 1):
            vals = row_values(ws, r)
            if vals:
                tail.append({'row': r, 'values': vals})
        sample_styles = []
        for r in range(1, min(ws.max_row, 30) + 1):
            for c in range(1, min(ws.max_column, 15) + 1):
                cell = ws.cell(r, c)
                if cell.value is not None and len(sample_styles) < 25:
                    sample_styles.append({
                        'cell': cell.coordinate, 'value': safe(cell.value), 'style_id': cell.style_id,
                        'number_format': cell.number_format, 'font': {'name': cell.font.name, 'size': cell.font.sz, 'bold': cell.font.b, 'color': safe(cell.font.color.rgb if cell.font.color and cell.font.color.type == 'rgb' else None)},
                        'fill': safe(cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == 'rgb' else None),
                        'alignment': {'horizontal': cell.alignment.horizontal, 'vertical': cell.alignment.vertical, 'wrap': cell.alignment.wrap_text},
                    })
        result['sheets'].append({
            'name': ws.title, 'state': ws.sheet_state, 'max_row': ws.max_row, 'max_column': ws.max_column,
            'freeze_panes': safe(ws.freeze_panes), 'merged_ranges': [str(x) for x in list(ws.merged_cells.ranges)[:100]],
            'tables': list(ws.tables.keys()), 'auto_filter': safe(ws.auto_filter.ref),
            'first_nonempty_rows': nonempty_rows, 'tail_nonempty_rows': tail,
            'sample_formulas': formula_cells, 'external_formulas': external_formula_cells,
            'sample_styles': sample_styles,
        })
    return result

out = Path('analysis_artifacts')
out.mkdir(exist_ok=True)
all_results = []
for src in SOURCES:
    print(f'Reading {src.name}', flush=True)
    all_results.append(summarize(src))
(out / 'structure_openpyxl.json').write_text(json.dumps(all_results, ensure_ascii=False, indent=2), encoding='utf-8')
print(out / 'structure_openpyxl.json')
